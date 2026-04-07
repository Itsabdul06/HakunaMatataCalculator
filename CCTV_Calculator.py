#!/usr/bin/env python3
"""
CCTV Master Calculator
Rewrite of KantechCalc with improved GUI.
Maintains all original functionality: camera entry, NVR management,
HDD pricing, auto/manual calculation, report export to Excel.
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import math, itertools, json, os
from datetime import datetime

# Try to import openpyxl for Excel export
try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ─────────────────────────── Persistence ───────────────────────────────────
DATA_FILE = "system_data.json"

DEFAULT_HDD_PRICES = {
    1: 87.0, 2: 131.0, 3: 145.0, 4: 239.0,
    6: 375.0, 8: 427.0, 10: 500.0, 12: 614.0,
    14: 1114.0, 18: 1291.0, 22: 1226.0, 24: 1568.0, 26: 2600.0,
}

DEFAULT_NVR_DATA = [
    {"Name": "1U RAID",        "SKU": "ADVER00N0NP16G", "CH": 32,  "MB": 50,   "Slots": 4,  "Price": 3750.0,  "mode": "RAID"},
    {"Name": "2U 64 Ch",       "SKU": "ADVER12R0N2H",   "CH": 64,  "MB": 300,  "Slots": 6,  "Price": 10416.7, "mode": "RAID"},
    {"Name": "2U 100 Ch",      "SKU": "ADVER00RN2J",    "CH": 100, "MB": 600,  "Slots": 8,  "Price": 11666.7, "mode": "RAID"},
    {"Name": "2U 128 Ch",      "SKU": "ADVER72R5N2H",   "CH": 128, "MB": 600,  "Slots": 12, "Price": 25000.0, "mode": "RAID"},
    {"Name": "2U Rack 175 Ch", "SKU": "ADVER02RDK",     "CH": 175, "MB": 1000, "Slots": 12, "Price": 13854.2, "mode": "RAID"},
    {"Name": "2U Rack 200 Ch", "SKU": "ADVER02RDK",     "CH": 200, "MB": 1500, "Slots": 12, "Price": 12812.5, "mode": "RAID"},
    {"Name": "Micro NVR",      "SKU": "ADVEM00N0NP8AH", "CH": 8,   "MB": 80,   "Slots": 1,  "Price": 1500.0,  "mode": "JBOD"},
    {"Name": "Desktop JBOD",   "SKU": "ADVED00N0N5H",   "CH": 50,  "MB": 200,  "Slots": 2,  "Price": 2291.7,  "mode": "JBOD"},
    {"Name": "2U 75 Ch",       "SKU": "ADVER00N0N2J",   "CH": 75,  "MB": 400,  "Slots": 4,  "Price": 5312.5,  "mode": "JBOD"},
    {"Name": "Holis 8 Ch",     "SKU": "HRN-08013P",     "CH": 8,   "MB": 160,  "Slots": 1,  "Price": 520.85,  "mode": "JBOD"},
    {"Name": "Holis 16 Ch",    "SKU": "HRN-16023P",     "CH": 16,  "MB": 320,  "Slots": 2,  "Price": 770.85,  "mode": "JBOD"},
]

# ─────────────────────────── Colors & Fonts ────────────────────────────────
BG       = "#0f1520"
SURFACE  = "#151d2e"
SURFACE2 = "#1a2540"
SURFACE3 = "#1f2d4a"
BORDER   = "#253046"
ACCENT   = "#00d4ff"
ACCENT_D = "#0099bb"
GREEN    = "#22d3a5"
GOLD     = "#f59e0b"
RED      = "#f87171"
TEXT     = "#e2e8f0"
TEXT2    = "#7a90b0"
TEXT3    = "#3d5070"
WHITE    = "#ffffff"

FONT_H1   = ("Segoe UI", 16, "bold")
FONT_H2   = ("Segoe UI", 11, "bold")
FONT_H3   = ("Segoe UI", 10, "bold")
FONT_BODY = ("Segoe UI", 9)
FONT_MONO = ("Consolas", 9)
FONT_BTN  = ("Segoe UI", 9, "bold")
FONT_LRGE = ("Segoe UI", 11, "bold")

# Constants
MAX_NVR_COMBOS = 5  # Maximum number of NVRs to consider in auto mode

# ─────────────────────────── Core Logic ────────────────────────────────────
def get_best_hdd(required_tb, slots, parity, price_dict):
    """Find the most cost-effective HDD configuration"""
    best_cost, best_cfg = float('inf'), None
    for cap in sorted(price_dict.keys()):
        price = price_dict[cap]
        min_drives = parity + 1
        data_req = math.ceil(required_tb / cap)
        total_drives = data_req + parity
        if total_drives > slots:
            continue
        total_drives = max(total_drives, min_drives)
        cost = total_drives * price
        if cost < best_cost:
            best_cost = cost
            best_cfg  = {"cap": cap, "qty": total_drives, "data": data_req, "cost": cost}
    return best_cfg

# ─────────────────────────── Widget Helpers ────────────────────────────────
def mk_frame(parent, bg=SURFACE, **kw):
    return tk.Frame(parent, bg=bg, **kw)

def mk_label(parent, text, font=FONT_BODY, fg=TEXT2, bg=SURFACE, anchor="w", **kw):
    return tk.Label(parent, text=text, font=font, fg=fg, bg=bg, anchor=anchor, **kw)

def mk_entry(parent, textvariable=None, width=12, font=FONT_MONO, **kw):
    defaults = dict(
        bg=SURFACE2, fg=TEXT, insertbackground=ACCENT,
        relief="flat", bd=0,
        highlightthickness=1, highlightbackground=BORDER,
        highlightcolor=ACCENT,
    )
    defaults.update(kw)
    e = tk.Entry(parent, textvariable=textvariable, width=width,
                 font=font, **defaults)
    return e

def mk_btn(parent, text, command, style="normal", **kw):
    colors = {
        "primary": (ACCENT,   "#000000", ACCENT_D),
        "danger":  (SURFACE2, RED,       SURFACE3),
        "ghost":   (SURFACE2, TEXT2,     SURFACE3),
        "success": (GREEN,    "#000000", "#18a87f"),
        "normal":  (SURFACE3, TEXT,      BORDER),
    }
    bg, fg, abg = colors.get(style, colors["normal"])
    return tk.Button(parent, text=text, command=command,
                     bg=bg, fg=fg, activebackground=abg, activeforeground=fg,
                     font=FONT_BTN, relief="flat", bd=0,
                     cursor="hand2", padx=10, pady=5, **kw)

def sep(parent, bg=BORDER, vertical=False):
    if vertical:
        return tk.Frame(parent, bg=bg, width=1)
    return tk.Frame(parent, bg=bg, height=1)

# ─────────────────────────── Application ───────────────────────────────────
class CCTVApp:
    def __init__(self, root):
        self.root = root
        self.root.title("CCTV Master Calculator")
        self.root.configure(bg=BG)
        self.root.geometry("1200x820")
        self.root.minsize(1000, 700)

        self.last_report = None
        self.last_calculation_result = None  # Store calculation result for Excel export
        self.hdd_ents    = {}
        self.nvr_price_entries = []
        self.progress_window = None

        self.load_all_data()
        self.setup_ui()
        self._apply_ttk_styles()

    # ── Data persistence ──────────────────────────────────────────────────
    def load_all_data(self):
        if os.path.exists(DATA_FILE):
            try:
                with open(DATA_FILE, "r") as f:
                    data = json.load(f)
                self.hdd_prices = {int(k): float(v) for k, v in data.get("hdd", {}).items()}
                self.nvr_list   = [dict(x) for x in data.get("nvr", [])]
                return
            except Exception:
                pass
        self.hdd_prices = dict(DEFAULT_HDD_PRICES)
        self.nvr_list   = [dict(n) for n in DEFAULT_NVR_DATA]

    def save_all_data(self):
        with open(DATA_FILE, "w") as f:
            json.dump({"hdd": self.hdd_prices, "nvr": self.nvr_list}, f, indent=2)

    # ── TTK styles ────────────────────────────────────────────────────────
    def _apply_ttk_styles(self):
        s = ttk.Style()
        s.theme_use("clam")
        s.configure("TNotebook", background=BG, borderwidth=0, tabmargins=0)
        s.configure("TNotebook.Tab",
                    background=SURFACE, foreground=TEXT2,
                    font=FONT_H3, padding=(16, 8),
                    borderwidth=0, focuscolor=BG)
        s.map("TNotebook.Tab",
              background=[("selected", SURFACE2), ("active", SURFACE3)],
              foreground=[("selected", ACCENT),   ("active", TEXT)])
        s.configure("Treeview",
                    background=SURFACE, foreground=TEXT,
                    fieldbackground=SURFACE, rowheight=24,
                    font=FONT_MONO, borderwidth=0)
        s.configure("Treeview.Heading",
                    background=SURFACE2, foreground=ACCENT,
                    font=FONT_H3, relief="flat", borderwidth=0)
        s.map("Treeview",
              background=[("selected", ACCENT_D)],
              foreground=[("selected", WHITE)])
        s.map("Treeview.Heading", relief=[("active", "flat")])
        s.configure("Vertical.TScrollbar",   background=BORDER, troughcolor=SURFACE, arrowcolor=TEXT3, borderwidth=0)
        s.configure("Horizontal.TScrollbar", background=BORDER, troughcolor=SURFACE, arrowcolor=TEXT3, borderwidth=0)
        s.configure("TCombobox",
                    fieldbackground=SURFACE2, background=SURFACE2,
                    foreground=TEXT, bordercolor=BORDER,
                    arrowcolor=ACCENT, selectbackground=SURFACE2,
                    selectforeground=TEXT)
        s.map("TCombobox",
              fieldbackground=[("readonly", SURFACE2)],
              foreground=[("readonly", TEXT)])

    # ── Build UI ──────────────────────────────────────────────────────────
    def setup_ui(self):
        hdr = mk_frame(self.root, bg=BG)
        hdr.pack(fill="x", padx=24, pady=(18, 0))
        mk_label(hdr, "CCTV Master Calculator", font=FONT_H1, fg=WHITE, bg=BG).pack(side="left")
        mk_label(hdr, "  v34.0", font=FONT_BODY, fg=TEXT3, bg=BG).pack(side="left", pady=(6, 0))
        sep(self.root).pack(fill="x", padx=24, pady=10)

        self.nb = ttk.Notebook(self.root, style="TNotebook")
        self.nb.pack(fill="both", expand=True, padx=20, pady=(0, 16))

        self.tabs = []
        for title in ["Cameras", "Calculate", "NVR Models", "HDD Prices"]:
            f = mk_frame(self.nb, bg=SURFACE2)
            self.nb.add(f, text=f"  {title}  ")
            self.tabs.append(f)

        self._build_cameras_tab(self.tabs[0])
        self._build_calc_tab(self.tabs[1])
        self._build_nvr_tab(self.tabs[2])
        self._build_hdd_tab(self.tabs[3])

    # ── Tab 1: Cameras ────────────────────────────────────────────────────
    def _build_cameras_tab(self, tab):
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(1, weight=1)

        inp = mk_frame(tab, bg=SURFACE)
        inp.grid(row=0, column=0, sticky="ew", padx=16, pady=14)

        mk_label(inp, "Add / Update Camera", font=FONT_H2, fg=ACCENT, bg=SURFACE).grid(
            row=0, column=0, columnspan=10, sticky="w", padx=14, pady=(10, 8))

        fields  = ["Name", "Count", "Mbps/cam", "Storage TB/cam"]
        defaults= ["Bullet Cam", "64", "4.0", "1.0"]
        self.ents = {}
        for col, (f, d) in enumerate(zip(fields, defaults)):
            mk_label(inp, f, bg=SURFACE, fg=TEXT2).grid(row=1, column=col*2, sticky="w", padx=(14 if col==0 else 8, 4))
            var = tk.StringVar(value=d)
            e = mk_entry(inp, textvariable=var, width=13)
            e.grid(row=1, column=col*2+1, padx=(0, 4), pady=(0, 10))
            self.ents[f] = var

        btn_f = mk_frame(inp, bg=SURFACE)
        btn_f.grid(row=1, column=len(fields)*2, padx=(8, 14), pady=(0, 10))
        mk_btn(btn_f, "Add / Update", self.save_camera, style="primary").pack(side="left", padx=(0, 6))
        mk_btn(btn_f, "Delete", self.delete_camera, style="danger").pack(side="left")

        sep(tab).grid(row=0, column=0, sticky="ew", padx=16)

        tree_f = mk_frame(tab, bg=SURFACE2)
        tree_f.grid(row=1, column=0, sticky="nsew", padx=16, pady=14)
        tree_f.columnconfigure(0, weight=1)
        tree_f.rowconfigure(0, weight=1)

        cols = ("Name", "Count", "Mbps/cam", "Storage TB/cam")
        self.tree = ttk.Treeview(tree_f, columns=cols, show="headings")
        widths = [260, 80, 100, 130]
        for c, w in zip(cols, widths):
            self.tree.heading(c, text=c)
            self.tree.column(c, width=w, anchor="center" if c != "Name" else "w")
        self.tree.tag_configure("odd",  background=SURFACE)
        self.tree.tag_configure("even", background=SURFACE2)

        vsb = ttk.Scrollbar(tree_f, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        self.tree.bind("<<TreeviewSelect>>", self._on_cam_select)

    def _on_cam_select(self, event):
        sel = self.tree.selection()
        if not sel: return
        vals = self.tree.item(sel[0])["values"]
        keys = ["Name", "Count", "Mbps/cam", "Storage TB/cam"]
        for k, v in zip(keys, vals):
            self.ents[k].set(str(v))

    def save_camera(self):
        try:
            name = self.ents["Name"].get().strip()
            if not name:
                raise ValueError("Camera name cannot be empty")
            
            count = self.ents["Count"].get().strip()
            if not count:
                raise ValueError("Count cannot be empty")
            
            mbps = self.ents["Mbps/cam"].get().strip()
            if not mbps:
                raise ValueError("Mbps/cam cannot be empty")
            
            storage = self.ents["Storage TB/cam"].get().strip()
            if not storage:
                raise ValueError("Storage TB/cam cannot be empty")
            
            float(count); float(mbps); float(storage)
            
            if float(count) <= 0:
                raise ValueError("Count must be positive")
            if float(mbps) <= 0:
                raise ValueError("Mbps/cam must be positive")
            if float(storage) <= 0:
                raise ValueError("Storage TB/cam must be positive")
                
        except ValueError as e:
            messagebox.showerror("Error", f"Invalid input: {e}")
            return
            
        sel = self.tree.selection()
        if sel:
            self.tree.delete(sel[0])
        tag = "even" if len(self.tree.get_children()) % 2 == 0 else "odd"
        self.tree.insert("", "end", values=(name, count, mbps, storage), tags=(tag,))
        self.refresh_nvr_dropdowns()

    def delete_camera(self):
        for s in self.tree.selection():
            self.tree.delete(s)
        self.refresh_nvr_dropdowns()

    # ── Tab 2: Calculate ──────────────────────────────────────────────────
    def _build_calc_tab(self, tab):
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(1, weight=1)

        ctrl = mk_frame(tab, bg=SURFACE)
        ctrl.grid(row=0, column=0, sticky="ew", padx=16, pady=14)

        mk_label(ctrl, "Calculation Settings", font=FONT_H2, fg=ACCENT, bg=SURFACE).pack(
            anchor="w", padx=14, pady=(10, 8))

        row = mk_frame(ctrl, bg=SURFACE)
        row.pack(fill="x", padx=14, pady=(0, 10))

        mk_label(row, "Mode:", bg=SURFACE, fg=TEXT2).grid(row=0, column=0, sticky="w", padx=(0, 6))
        self.auto_mode = tk.StringVar(value="AUTO")
        for val, lbl in (("AUTO", "Auto (find best NVR combo)"), ("MANUAL", "Manual (choose NVR below)")):
            rb = tk.Radiobutton(row, text=lbl, variable=self.auto_mode, value=val,
                                bg=SURFACE, fg=TEXT2, selectcolor=SURFACE2,
                                activebackground=SURFACE, activeforeground=TEXT,
                                font=FONT_BODY, command=self._on_mode_change)
            rb.grid(row=0, column=(1 if val=="AUTO" else 2), padx=(0, 16))

        mk_label(row, "RAID Level:", bg=SURFACE, fg=TEXT2).grid(row=0, column=3, sticky="w", padx=(16, 6))
        self.raid_var = tk.StringVar(value="RAID 5")
        cb_raid = ttk.Combobox(row, textvariable=self.raid_var, width=10,
                               state="readonly", values=["JBOD", "RAID 5", "RAID 6"])
        cb_raid.grid(row=0, column=4, padx=(0, 16))

        self.manual_frame = mk_frame(ctrl, bg=SURFACE)
        self.manual_frame.pack(fill="x", padx=14, pady=(0, 10))
        mk_label(self.manual_frame, "Manual NVR:", bg=SURFACE, fg=TEXT2).grid(row=0, column=0, padx=(0, 6))
        self.manual_combos = []
        for i in range(6):
            var = tk.StringVar(value="None")
            cb  = ttk.Combobox(self.manual_frame, textvariable=var, width=16,
                               state="readonly", values=["None"])
            cb.grid(row=0, column=i+1, padx=4)
            self.manual_combos.append(cb)
        self.manual_frame.grid_remove()

        btn_row = mk_frame(ctrl, bg=SURFACE)
        btn_row.pack(fill="x", padx=14, pady=(0, 12))
        mk_btn(btn_row, "⚡  Run Calculation", self.run_logic, style="primary").pack(side="left", padx=(0, 10))
        mk_btn(btn_row, "Export to Excel", self.export_to_excel, style="success").pack(side="left", padx=(0, 10))
        self.calc_status = mk_label(btn_row, "", fg=TEXT2, bg=SURFACE, font=FONT_BODY)
        self.calc_status.pack(side="left", padx=16)

        sep(tab).grid(row=0, column=0, sticky="ew", padx=16)

        res_f = mk_frame(tab, bg=SURFACE2)
        res_f.grid(row=1, column=0, sticky="nsew", padx=16, pady=14)
        res_f.columnconfigure(0, weight=1)
        res_f.rowconfigure(1, weight=1)

        mk_label(res_f, "Results", font=FONT_H2, fg=WHITE, bg=SURFACE2).grid(
            row=0, column=0, columnspan=2, sticky="w", padx=14, pady=(10, 6))

        self.res_txt = tk.Text(res_f, bg=SURFACE, fg=TEXT, font=FONT_MONO,
                               relief="flat", bd=0, state="disabled",
                               highlightthickness=0, wrap="none",
                               padx=14, pady=10)
        vsb2 = ttk.Scrollbar(res_f, orient="vertical", command=self.res_txt.yview)
        hsb2 = ttk.Scrollbar(res_f, orient="horizontal", command=self.res_txt.xview)
        self.res_txt.configure(yscrollcommand=vsb2.set, xscrollcommand=hsb2.set)
        self.res_txt.grid(row=1, column=0, sticky="nsew")
        vsb2.grid(row=1, column=1, sticky="ns")
        hsb2.grid(row=2, column=0, sticky="ew")

        self.res_txt.tag_configure("header",  foreground=ACCENT,   font=("Consolas", 9, "bold"))
        self.res_txt.tag_configure("best",    foreground=GREEN,    font=("Consolas", 9, "bold"))
        self.res_txt.tag_configure("label",   foreground=TEXT2)
        self.res_txt.tag_configure("value",   foreground=TEXT)
        self.res_txt.tag_configure("divider", foreground=TEXT3)
        self.res_txt.tag_configure("cost",    foreground=GOLD,     font=("Consolas", 9, "bold"))
        self.res_txt.tag_configure("error",   foreground=RED)

        self.refresh_nvr_dropdowns()
        self._on_mode_change()

    def _on_mode_change(self):
        if self.auto_mode.get() == "MANUAL":
            self.manual_frame.grid()
        else:
            self.manual_frame.grid_remove()

    def refresh_nvr_dropdowns(self):
        names = ["None"] + [n["Name"] for n in self.nvr_list]
        for combo in self.manual_combos:
            current = combo.get()
            combo['values'] = names
            if current not in names:
                combo.set("None")

    def show_progress(self):
        if self.progress_window and self.progress_window.winfo_exists():
            return
            
        self.progress_window = tk.Toplevel(self.root)
        self.progress_window.title("Calculating...")
        self.progress_window.configure(bg=SURFACE)
        self.progress_window.geometry("300x100")
        self.progress_window.transient(self.root)
        self.progress_window.grab_set()
        
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - 150
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - 50
        self.progress_window.geometry(f"300x100+{x}+{y}")
        
        mk_label(self.progress_window, "Analyzing configurations...", 
                font=FONT_H2, fg=ACCENT, bg=SURFACE).pack(pady=(20, 10))
        
        self.progress_bar = ttk.Progressbar(self.progress_window, mode='indeterminate')
        self.progress_bar.pack(padx=20, pady=10, fill='x')
        self.progress_bar.start(10)
        
        self.root.update()

    def hide_progress(self):
        if self.progress_window and self.progress_window.winfo_exists():
            self.progress_window.destroy()
        self.progress_window = None

    # ── Tab 3: NVR Models ─────────────────────────────────────────────────
    def _build_nvr_tab(self, tab):
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(1, weight=1)

        add_f = mk_frame(tab, bg=SURFACE)
        add_f.grid(row=0, column=0, sticky="ew", padx=16, pady=14)
        mk_label(add_f, "Add New NVR Model", font=FONT_H2, fg=ACCENT, bg=SURFACE).grid(
            row=0, column=0, columnspan=12, sticky="w", padx=14, pady=(10, 8))

        self.nf = {}
        fields = [("Name", 14), ("SKU", 14), ("CH", 6), ("MB", 6), ("Slots", 6), ("Price", 8)]
        for col, (f, w) in enumerate(fields):
            mk_label(add_f, f, bg=SURFACE, fg=TEXT2).grid(row=1, column=col*2, sticky="w", padx=(14 if col==0 else 6, 3))
            var = tk.StringVar()
            e = mk_entry(add_f, textvariable=var, width=w)
            e.grid(row=1, column=col*2+1, padx=(0, 2), pady=(0, 10))
            self.nf[f] = var

        self.na = tk.StringVar(value="RAID")
        mk_label(add_f, "RAID/JBOD", bg=SURFACE, fg=TEXT2).grid(row=1, column=12, sticky="w", padx=(6, 3))
        ttk.Combobox(add_f, textvariable=self.na, width=7,
                     state="readonly", values=["RAID", "JBOD"]).grid(row=1, column=13, padx=(0, 6), pady=(0, 10))
        mk_btn(add_f, "ADD TO DATABASE", self.add_new_nvr, style="primary").grid(
            row=1, column=14, padx=(6, 14), pady=(0, 10))

        sep(tab).grid(row=0, column=0, sticky="ew", padx=16)

        list_outer = mk_frame(tab, bg=SURFACE2)
        list_outer.grid(row=1, column=0, sticky="nsew", padx=16, pady=14)
        list_outer.columnconfigure(0, weight=1)
        list_outer.rowconfigure(0, weight=1)

        canvas = tk.Canvas(list_outer, bg=SURFACE2, highlightthickness=0)
        vsb    = ttk.Scrollbar(list_outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        self.nvr_frame = mk_frame(canvas, bg=SURFACE2)
        self.nvr_canvas_win = canvas.create_window((0, 0), window=self.nvr_frame, anchor="nw")

        def _on_resize(e):
            canvas.itemconfig(self.nvr_canvas_win, width=e.width)
        canvas.bind("<Configure>", _on_resize)
        self.nvr_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        self.nvr_canvas = canvas
        self.nvr_price_entries = []

        hdr = mk_frame(self.nvr_frame, bg=SURFACE3)
        hdr.pack(fill="x", pady=(0, 2))
        for txt, w in [("Name", 160), ("SKU", 140), ("Channels", 70), ("Max MB/s", 75),
                       ("HDD Slots", 75), ("Price ($)", 90), ("Mode", 60), ("", 100)]:
            mk_label(hdr, txt, font=FONT_H3, fg=ACCENT, bg=SURFACE3, width=w//8, anchor="w").pack(
                side="left", padx=8, pady=6)

        self.refresh_nvr_list_tab()

    def refresh_nvr_list_tab(self):
        for w in self.nvr_frame.winfo_children():
            if w != self.nvr_frame.winfo_children()[0]:
                w.destroy()
        self.nvr_price_entries = []

        for i, n in enumerate(self.nvr_list):
            row_bg = SURFACE if i % 2 == 0 else SURFACE2
            row = mk_frame(self.nvr_frame, bg=row_bg)
            row.pack(fill="x", pady=1)

            mk_label(row, n["Name"],         bg=row_bg, fg=TEXT,  width=20).pack(side="left", padx=(12,4), pady=4)
            mk_label(row, n["SKU"],          bg=row_bg, fg=TEXT2, font=FONT_MONO, width=16).pack(side="left", padx=4)
            mk_label(row, str(n["CH"]),      bg=row_bg, fg=TEXT,  width=9,  anchor="center").pack(side="left", padx=4)
            mk_label(row, str(n["MB"]),      bg=row_bg, fg=TEXT,  width=9,  anchor="center").pack(side="left", padx=4)
            mk_label(row, str(n["Slots"]),   bg=row_bg, fg=TEXT,  width=9,  anchor="center").pack(side="left", padx=4)

            price_var = tk.StringVar(value=f"{n['Price']:.2f}")
            e = mk_entry(row, textvariable=price_var, width=10, bg=row_bg)
            e.pack(side="left", padx=4)
            self.nvr_price_entries.append(price_var)

            mk_label(row, n.get("mode", "RAID"), bg=row_bg, fg=GOLD if n.get("mode")=="RAID" else ACCENT,
                     width=7, anchor="center").pack(side="left", padx=4)

            mk_btn(row, "Delete", lambda idx=i: self.delete_nvr(idx), style="danger").pack(
                side="right", padx=(4, 12))

        save_row = mk_frame(self.nvr_frame, bg=SURFACE2)
        save_row.pack(fill="x", pady=8, padx=12)
        mk_btn(save_row, "Save All Price Updates", self.save_nvr_prices, style="success").pack(side="left")

    def add_new_nvr(self):
        try:
            name = self.nf["Name"].get().strip()
            if not name:
                raise ValueError("Name required")
            
            sku = self.nf["SKU"].get().strip()
            if not sku:
                raise ValueError("SKU required")
            
            ch_str = self.nf["CH"].get().strip()
            if not ch_str:
                raise ValueError("Channels required")
            
            mb_str = self.nf["MB"].get().strip()
            if not mb_str:
                raise ValueError("Max MB/s required")
            
            slots_str = self.nf["Slots"].get().strip()
            if not slots_str:
                raise ValueError("HDD Slots required")
            
            price_str = self.nf["Price"].get().strip()
            if not price_str:
                raise ValueError("Price required")
            
            row = {
                "Name":  name,
                "SKU":   sku,
                "CH":    int(ch_str),
                "MB":    int(mb_str),
                "Slots": int(slots_str),
                "Price": float(price_str),
                "mode":  self.na.get(),
            }
            
            if row["CH"] <= 0:
                raise ValueError("Channels must be positive")
            if row["MB"] <= 0:
                raise ValueError("Max MB/s must be positive")
            if row["Slots"] <= 0:
                raise ValueError("HDD Slots must be positive")
            if row["Price"] <= 0:
                raise ValueError("Price must be positive")
                
            self.nvr_list.append(row)
            self.save_all_data()
            self.refresh_nvr_dropdowns()
            self.refresh_nvr_list_tab()
            messagebox.showinfo("Success", "NVR Added.")
        except Exception as e:
            messagebox.showerror("Error", f"Invalid input: {e}")

    def save_nvr_prices(self):
        for i, var in enumerate(self.nvr_price_entries):
            try:
                self.nvr_list[i]["Price"] = float(var.get())
            except ValueError:
                pass
        self.save_all_data()
        messagebox.showinfo("Saved", "NVR Prices Updated.")

    def delete_nvr(self, idx):
        if messagebox.askyesno("Confirm", "Delete this model?"):
            self.nvr_list.pop(idx)
            self.save_all_data()
            self.refresh_nvr_dropdowns()
            self.refresh_nvr_list_tab()

    # ── Tab 4: HDD Prices ─────────────────────────────────────────────────
    def _build_hdd_tab(self, tab):
        tab.columnconfigure(0, weight=1)

        outer = mk_frame(tab, bg=SURFACE)
        outer.grid(row=0, column=0, sticky="nsew", padx=16, pady=14)

        mk_label(outer, "Hard Drive Prices  (EGP per drive)", font=FONT_H2, fg=ACCENT, bg=SURFACE).pack(
            anchor="w", padx=14, pady=(12, 10))

        grid = mk_frame(outer, bg=SURFACE)
        grid.pack(fill="x", padx=14, pady=(0, 10))

        self.hdd_ents = {}
        for i, cap in enumerate(sorted(self.hdd_prices.keys())):
            col, row = (i % 4) * 3, i // 4
            mk_label(grid, f"{cap} TB", fg=TEXT2, bg=SURFACE, width=6).grid(
                row=row, column=col, sticky="w", padx=(0, 4), pady=5)
            var = tk.StringVar(value=f"{self.hdd_prices[cap]:.2f}")
            e = mk_entry(grid, textvariable=var, width=10)
            e.grid(row=row, column=col+1, padx=(0, 24), pady=5)
            self.hdd_ents[cap] = var

        btn_row = mk_frame(outer, bg=SURFACE)
        btn_row.pack(anchor="w", padx=14, pady=(6, 14))
        mk_btn(btn_row, "Save HDD Prices", self.save_hdds, style="success").pack(side="left")

    def save_hdds(self):
        for cap, var in self.hdd_ents.items():
            try:
                price = float(var.get())
                if price <= 0:
                    raise ValueError("Price must be positive")
                self.hdd_prices[cap] = price
            except ValueError as e:
                messagebox.showerror("Error", f"Invalid price for {cap}TB: {e}")
                return
        self.save_all_data()
        messagebox.showinfo("Saved", "HDD Prices Updated.")

    # ── Calculation engine ────────────────────────────────────────────────
    def calculate_engine(self, cams, hw_cfg, split_ratio=None):
        n_units = len(hw_cfg)
        u_list  = []
        cur_c   = 0

        for i, hw in enumerate(hw_cfg):
            if split_ratio is None:
                take = math.ceil(len(cams) / n_units) if i < n_units-1 else len(cams) - cur_c
            else:
                take = math.ceil(len(cams) * split_ratio[i])
                take = min(take, len(cams) - cur_c)

            u_brk = cams[cur_c:cur_c + take]
            cur_c += take

            u_mb  = sum(int(c[1]) * float(c[2]) for c in u_brk)
            u_tb  = sum(int(c[1]) * float(c[3]) for c in u_brk)
            u_c   = sum(int(c[1]) for c in u_brk)

            if u_c > hw["CH"]:
                return None

            raid = self.raid_var.get()
            parity = 0 if raid == "JBOD" else (1 if raid == "RAID 5" else 2)
            mode_str = raid

            hd = get_best_hdd(u_tb, hw["Slots"], parity, self.hdd_prices)
            if hd is None:
                return None

            if u_mb > hw["MB"]:
                return None

            cam_breakdown = {c[0]: int(c[1]) for c in u_brk}
            u_list.append({
                "name": hw["Name"], "m": hw, "mode": mode_str,
                "mb": u_mb, "load": u_mb / hw["MB"] * 100 if hw["MB"] > 0 else 0,
                "c_total": u_c, "cam_breakdown": cam_breakdown,
                "qty": hd["qty"], "cap": hd["cap"],
                "total_tb": hd["qty"] * hd["cap"],
                "cost": hw["Price"] + hd["cost"], "h": hd,
            })

        return u_list

    def run_logic(self):
        cams = [self.tree.item(i)["values"] for i in self.tree.get_children()]
        if not cams:
            messagebox.showwarning("Warning", "Add cameras first.")
            return

        for cam in cams:
            try:
                if not cam[0].strip():
                    raise ValueError("Camera name cannot be empty")
                count = int(cam[1])
                mbps = float(cam[2])
                storage = float(cam[3])
                if count <= 0 or mbps <= 0 or storage <= 0:
                    raise ValueError("All camera values must be positive")
            except (ValueError, IndexError) as e:
                messagebox.showerror("Error", f"Invalid camera data: {cam[0] if cam else 'Unknown'}\n{e}")
                return

        self.calc_status.config(text="Calculating...", fg=GOLD)
        self.show_progress()
        self.root.update()

        try:
            if self.auto_mode.get() == "AUTO":
                pool = [n for n in self.nvr_list if (
                    n.get("mode", "RAID") == "JBOD" and self.raid_var.get() == "JBOD" or
                    n.get("mode", "RAID") == "RAID" and self.raid_var.get() != "JBOD"
                )]
                if not pool:
                    pool = list(self.nvr_list)

                total_cams = sum(int(c[1]) for c in cams)
                best_cfg, best_cost = None, float("inf")
                max_nvrs = min(MAX_NVR_COMBOS, len(pool))
                
                for n_u in range(1, max_nvrs + 1):
                    for combo in itertools.combinations_with_replacement(pool, n_u):
                        hw_c = list(combo)
                        max_ch = sum(n["CH"] for n in hw_c)
                        if max_ch < total_cams:
                            continue
                        res = self.calculate_engine(cams, hw_c, None)
                        if res is None:
                            continue
                        cost = sum(x["cost"] for x in res)
                        if cost < best_cost:
                            best_cost = cost
                            best_cfg  = res
                txt = best_cfg
            else:
                active_hw = []
                for combo in self.manual_combos:
                    nvr_name = combo.get()
                    if nvr_name != "None":
                        nvr = next((n for n in self.nvr_list if n["Name"] == nvr_name), None)
                        if nvr:
                            active_hw.append(nvr)
                
                if not active_hw:
                    messagebox.showwarning("Warning", "Select at least one NVR.")
                    self.calc_status.config(text="", fg=TEXT2)
                    self.hide_progress()
                    return
                
                txt = self.calculate_engine(cams, active_hw, None)

            if not txt:
                self._show_result_error("ERROR: No valid configuration found.\n\nPossible reasons:\n• HDD sizes cannot meet storage requirements\n• NVR channel/slot limits exceeded\n• No compatible NVRs available for selected RAID mode")
                self.calc_status.config(text="No solution found", fg=RED)
                self.hide_progress()
                return

            self.last_calculation_result = {
                "cameras": cams,
                "nvr_config": txt,
                "raid_mode": self.raid_var.get()
            }
            
            self.generate_detailed_report(txt)
            total_cost = sum(x["cost"] for x in txt)
            self.calc_status.config(text=f"Done — Total: ${total_cost:,.2f}", fg=GREEN)

        except Exception as e:
            self._show_result_error(f"ERROR: {str(e)}")
            self.calc_status.config(text="Error", fg=RED)
        finally:
            self.hide_progress()

    def _show_result_error(self, msg):
        self.res_txt.config(state="normal")
        self.res_txt.delete("1.0", "end")
        self.res_txt.insert("end", msg, "error")
        self.res_txt.config(state="disabled")

    def generate_detailed_report(self, cfg):
        now    = datetime.now().strftime("%Y-%m-%d %H:%M")
        total  = sum(u["cost"] for u in cfg)
        lines  = []

        def write(text, tag="value"):
            lines.append((text, tag))

        write("=" * 72 + "\n", "divider")
        write(f" CCTV DESIGN REPORT  —  {now}\n", "header")
        write(f" SYSTEM TOTAL: ${total:,.2f}\n", "cost")
        write("=" * 72 + "\n", "divider")

        for i, u in enumerate(cfg, 1):
            write(f"\nUNIT #{i}: {u['name']}\n", "best")
            write("-" * 50 + "\n", "divider")
            write(f"  Mode:     ", "label"); write(f"{u['mode']}\n", "value")
            write(f"  Load:     ", "label"); write(f"{u['mb']:.1f} Mbps  ({u['load']:.1f}% of {u['m']['MB']} MB/s capacity)\n", "value")
            write(f"  Cameras:  ", "label"); write(f"{u['c_total']} total  ", "value")
            if u["cam_breakdown"]:
                parts = ",  ".join(f"{n}: {c}" for n, c in u["cam_breakdown"].items())
                write(f"({parts})\n", "value")
            else:
                write("\n", "value")
            write(f"  Storage:  ", "label")
            write(f"{u['qty']} × {u['cap']} TB  = {u['total_tb']} TB  ", "value")
            write(f"(usable: {u['h']['data']} × {u['cap']} = {u['h']['data']*u['cap']} TB)\n", "label")
            write(f"  Cost:     ", "label")
            write(f"NVR ${u['m']['Price']:,.2f}  +  HDD ${u['h']['cost']:,.2f}  =  ${u['cost']:,.2f}\n", "cost")

        write("\n" + "=" * 72 + "\n", "divider")
        write(f" GRAND TOTAL:  ${total:,.2f}\n", "cost")
        write("=" * 72 + "\n", "divider")

        self.res_txt.config(state="normal")
        self.res_txt.delete("1.0", "end")
        for text, tag in lines:
            self.res_txt.insert("end", text, tag)
        self.res_txt.config(state="disabled")

        self.last_report = "".join(t for t, _ in lines)
        self.nb.select(self.tabs[1])

    # ── Excel Export Function ─────────────────────────────────────────────
    def export_to_excel(self):
        """Export calculation results to Excel template"""
        
        # Check if we have calculation results
        if not self.last_calculation_result:
            messagebox.showwarning("Warning", "Run a calculation first before exporting!")
            return
        
        # Check if openpyxl is available
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Error", 
                "Excel export requires openpyxl library.\n\n"
                "Please install it using:\npip install openpyxl")
            return
        
        # Ask for template file location
        template_file = filedialog.askopenfilename(
            title="Select Excel Template",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            defaultextension=".xlsx"
        )
        
        if not template_file:
            return
        
        try:
            # Load the template
            wb = load_workbook(template_file)
            
            # Check if "offer" sheet exists
            if "offer" not in wb.sheetnames:
                messagebox.showerror("Error", "Sheet 'offer' not found in the template!")
                return
            
            ws = wb["offer"]
            
            # Get the Header 1 style from the template (if it exists)
            header_style = None
            if "Header 1" in wb._named_styles:
                header_style = wb._named_styles["Header 1"]
            
            # Start from row 9
            current_row = 9
            
            cameras = self.last_calculation_result["cameras"]
            nvr_config = self.last_calculation_result["nvr_config"]
            
            # Group NVRs by SKU and HDD config
            nvr_groups = {}
            for unit in nvr_config:
                sku = unit["m"]["SKU"]
                hdd_cap = unit["cap"]
                hdd_qty = unit["qty"]
                key = (sku, hdd_cap, hdd_qty)
                if key not in nvr_groups:
                    nvr_groups[key] = {
                        "sku": sku,
                        "nvr_name": unit["name"],
                        "hdd_cap": hdd_cap,
                        "hdd_qty": hdd_qty,
                        "count": 1,
                        "price": unit["m"]["Price"]
                    }
                else:
                    nvr_groups[key]["count"] += 1
            
            # Prepare data rows
            excel_rows = []
            
            # Add Camera Header
            excel_rows.append(("Cameras", "", "", "", "", "header"))
            
            # Add each camera type and its CAMLIC
            for cam in cameras:
                cam_name = cam[0]
                cam_qty = int(cam[1])
                
                # Camera row
                excel_rows.append((cam_name, cam_qty, "", "CCTV", "Camera", "data"))
                # CAMLIC row (immediately after each camera)
                excel_rows.append(("CAMLIC", 1, "ch", "CCTV", "Software", "data"))
            
            # Add NVRs Header
            excel_rows.append(("NVRs", "", "", "", "", "header"))
            
            # Add NVR groups and their HDDs
            for key, group in nvr_groups.items():
                # NVR row
                excel_rows.append((group["sku"], group["count"], "", "CCTV", "NVR", "data"))
                # HDD row for this NVR
                hdd_part_no = f"{group['hdd_cap']}TB HDD"
                excel_rows.append((hdd_part_no, group["hdd_qty"], "ch", "CCTV", "HDD", "data"))
            
            # Add VMS Header
            excel_rows.append(("VMS", "", "", "", "", "header"))
            
            # Add VMS row (Sys is blank, not "ch")
            excel_rows.append(("VMS", 1, "", "CCTV", "Software", "data"))
            
            # Write to Excel (overwrite only columns F, H, K, L, M)
            # Column mapping: F=5, H=7, K=10, L=11, M=12 (1-indexed)
            for row_data in excel_rows:
                part_no, qty, sys, solution, category, row_type = row_data
                
                # Write to columns F, H, K, L, M only
                if part_no:
                    ws.cell(row=current_row, column=6, value=part_no)  # Column F
                if qty:
                    ws.cell(row=current_row, column=8, value=qty)      # Column H
                if sys:
                    ws.cell(row=current_row, column=11, value=sys)     # Column K
                if solution:
                    ws.cell(row=current_row, column=12, value=solution) # Column L
                if category:
                    ws.cell(row=current_row, column=13, value=category) # Column M
                
                # Apply Header 1 style if this is a header row
                if row_type == "header" and header_style:
                    for col in [6, 8, 11, 12, 13]:  # Columns F, H, K, L, M
                        cell = ws.cell(row=current_row, column=col)
                        cell.style = "Header 1"
                
                current_row += 1
            
            # Save the file (overwrite)
            wb.save(template_file)
            
            messagebox.showinfo("Success", 
                f"Excel file has been updated successfully!\n\n"
                f"File: {os.path.basename(template_file)}\n"
                f"Sheet: offer\n"
                f"Rows added: {len(excel_rows)}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export to Excel:\n\n{str(e)}")

# ─────────────────────────── Entry Point ───────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("1200x820")
    app = CCTVApp(root)
    root.mainloop()
